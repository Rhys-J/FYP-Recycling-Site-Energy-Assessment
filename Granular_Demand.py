import datetime
import openpyxl
from random import random
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import math

workbook = openpyxl.load_workbook("Sensor_Data/Granular Current Reading.xlsx")

#----------------------SHREDDING-----------------------------------


SH_sheet = workbook["Shredder Line"]
SH_readings = []
SH_timings = []

for row in SH_sheet.iter_rows(4, 1000, 5, 6, True):
    SH_readings.append(row[1] * 400 * math.sqrt(3) * 0.8 / 1000)
    SH_timings.append(row[0])

plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
plt.plot(SH_timings, SH_readings)
plt.gcf().autofmt_xdate()
plt.ylim([0, max(SH_readings)+30])
plt.title("Shredder Line Granular Demand")
plt.xlabel("Time")
plt.ylabel("kW")
plt.show()

#----------------------GRANULATING-----------------------------------


GL_sheet = workbook["Granulation Line"]
GL_readings = []
GL_timings = []

for row in GL_sheet.iter_rows(4, 1000, 2, 3, True):
    GL_readings.append(row[1] * 400 * math.sqrt(3) * 0.8 / 1000)
    GL_timings.append(row[0])

plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
plt.plot(GL_timings, GL_readings)
plt.gcf().autofmt_xdate()
plt.ylim([0, max(GL_readings)+30])
plt.title("Granulation Line Granular Demand")
plt.xlabel("Time")
plt.ylabel("kW")
plt.show()

#----------------------TYREWIRE-----------------------------------


TW_sheet = workbook["TyreWire Line"]
TW_readings = []
TW_timings = []

for row in TW_sheet.iter_rows(4, 950, 2, 3, True):
    TW_readings.append(row[1] * 400 * math.sqrt(3) * 0.8 / 1000)
    TW_timings.append(row[0])

plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
plt.plot(TW_timings, TW_readings)
plt.gcf().autofmt_xdate()
plt.ylim([0, max(TW_readings)+30])
plt.title("Tyrewire Line Granular Demand")
plt.xlabel("Time")
plt.ylabel("kW")
plt.show()

#----------------------GL + TW-----------------------------------


GT_sheet = workbook["GL+TW"]
GT_readings = []
GT_timings = []

for row in GT_sheet.iter_rows(4, 950, 5, 6, True):
    GT_readings.append(row[1] * 400 * math.sqrt(3) * 0.8 / 1000)
    GT_timings.append(row[0])

plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
plt.plot(GT_timings, GT_readings)
plt.gcf().autofmt_xdate()
plt.ylim([0, max(GT_readings)+30])
plt.title("Granulation & Tyrewire Granular Demand")
plt.xlabel("Time")
plt.ylabel("kW")
plt.show()
